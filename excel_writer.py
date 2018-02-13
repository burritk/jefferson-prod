from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.workbook.views import BookView

def write_to_excel(input_txt, output_xlsx):
    with open(input_txt + '.txt', 'r') as input_file:
        with open('temp.txt', 'r+') as temp_file:
            row = 1
            for line in input_file:
                final_string = ''
                is_biz = False
                for index, field in enumerate(line.split('"')):
                    field = field.strip()
                    if index == 1:
                        # field is owner
                        owner = field.split(',')
                        if any(biz in field for biz in [' LLC', ' INC']):
                            field += '""'
                            is_biz = True
                        elif len(field.split(',')) > 1:
                            field = owner[1] + '"' #first name
                            field += owner[0] + '"' #last name
                        else:
                            field += '""'
                    elif index == 9:
                        # field is state city zip
                        address = field.split()
                        if len(address) > 2 and len(address[-2]) == 2:
                            field = ''
                            for i in range(0, len(address) - 2):  # city
                                field += address[i] + " "
                            field += '"' + address[-2] + '"'  # state
                            field += address[-1] + '"'  # zip
                        else:
                            field += '"""'
                    else:
                        field += '"'
                    final_string += field

                if is_biz:
                    final_string += 'BIZ'
                else:
                    final_string += ''

                temp_file.write(final_string + "\n")
                row += 1


    with open('temp.txt', 'r') as input_file:
        wb = Workbook()
        ws = wb.active
        row = 1

        for line in input_file:
            for index, field in enumerate(line.split('"')):
                print(row)
                ws[get_column_letter(index+1) + str(row)] = field
            row += 1
        wb.save(output_xlsx + '.xlsx')

write_to_excel('onethroughnineteen', 'final_ver')